import os
import openpyxl
from datetime import datetime, timedelta
import smtplib
from email.message import EmailMessage
from fpdf import FPDF

# 📁 Constants
TEMPLATE_PATH = "CB-Invoice-20260101.xlsx"
RECIPIENT_EMAIL = "recipient@example.com"  # Replace with actual recipient

# 📧 Credentials from environment (GitHub Secrets)
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")

# 📅 Month names
MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

# 📤 Email sender function
def send_email_with_attachment(to_email, subject, body, attachment_path):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = EMAIL_USER
    msg["To"] = to_email
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        file_data = f.read()
        msg.add_attachment(file_data, maintype="application", subtype="pdf", filename=os.path.basename(attachment_path))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_USER, EMAIL_PASS)
        smtp.send_message(msg)

# 🧾 PDF converter (basic text-based using FPDF)
def convert_to_pdf(invoice_data, pdf_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for line in invoice_data:
        pdf.cell(200, 10, txt=line, ln=True)

    pdf.output(pdf_path)

# 🚀 Main loop
for month in range(1, 13):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Update fields
    invoice_number = f"INVOICE #{20260000 + month * 100 + 1}"
    invoice_date = (datetime(2026, month, 1) + timedelta(days=31)).replace(day=1)
    description = f"Rental for {MONTH_NAMES[month - 1]} 2026"

    ws["B1"] = invoice_number
    ws["B6"] = invoice_date.strftime("%d/%m/%Y")
    ws["B13"] = description

    # Save Excel
    excel_filename = f"Invoice_{2026}{month:02d}01.xlsx"
    wb.save(excel_filename)

    # Extract data for PDF (simplified)
    invoice_data = [
        f"{invoice_number}",
        f"Date: {invoice_date.strftime('%d/%m/%Y')}",
        f"Description: {description}"
    ]
    pdf_filename = excel_filename.replace(".xlsx", ".pdf")
    convert_to_pdf(invoice_data, pdf_filename)

    # Send email
    subject = f"Invoice for {MONTH_NAMES[month - 1]} 2026"
    body = f"Dear Customer,\n\nPlease find attached your invoice for {MONTH_NAMES[month - 1]} 2026.\n\nBest regards,\nThirupathi"
    send_email_with_attachment(RECIPIENT_EMAIL, subject, body, pdf_filename)

print("✅ All invoices generated, converted to PDF, and emailed.")
